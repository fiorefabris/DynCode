import numpy as np
import pandas as pd
from openpyxl import load_workbook
from itertools import chain


def percentiles(conc,conc_order, save=False, *argv, **kwargs):
    """"Calculates the percentiles of each *argv
    if save=True, save the results on the "Percentiles" sheet of the data_chart.

    Returns:
        - a dictionary of df in wich each key is an *argv. Each df has the concentration as rows and the percentiles as cols.
    """
    #atencion agregue conc_order como variable
    amp_values = [amp_values for amp_values in argv]
    percentiles_dict = {}
    Index = [0, 3, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85,87, 90,91,92,93,94, 95,96, 97,98,99, 100]

    for amp_value in amp_values:
        Percentiles = pd.DataFrame(index=conc)
        for i in Index:
            Aux = []
            for C in conc_order:
                Aux.append(np.percentile(conc[C][amp_value], i))
            Percentiles[str(i)] = Aux
        percentiles_dict[amp_value] = Percentiles

    if save:
        perc_df = pd.concat([percentiles_dict[amp_value] for amp_value in percentiles_dict])
        perc_df['Amp Value'] = list(
            chain.from_iterable([[amp_value] * len(percentiles_dict[amp_value]) for amp_value in percentiles_dict]))

        save_folder = kwargs.get('save_folder', None)
        save_name = kwargs.get('save_name', None)
        book = load_workbook(str(save_folder) + str(save_name)+ '.xlsx')
        writer = pd.ExcelWriter(str(save_folder) + str(save_name)+ '.xlsx', engine='openpyxl')
        writer.book = book
        perc_df.to_excel(writer, sheet_name='Percentiles')
        writer.save()

    return percentiles_dict


def make_statistics(conc, save=False, *argv, **kwargs):
    """"Calculates the statistics of each *argv: ['amp_value', 'Concentration', 'cell', 'Mean', 'Var', 'Len', 'Median']

    note: the len value is expressed in hours. The 15 frames despicted at the beggining are considered.

    if save=True, save the results on the "Statistics"sheet of the data_chart.

    Returns:
        - a dictionary of df in wich each key is an *argv. Each df has the concentration as rows and the percentiles as cols.
    """
    doc_cols = ['value', 'Concentration', 'cell', 'Mean', 'Var', 'Len', 'Median']
    doc = pd.DataFrame(columns=doc_cols)

    amp_values = [amp_values for amp_values in argv]

    for amp_value in amp_values:
        for C in conc:
            df = conc[C]

            for cells, data in df.groupby(level='cell'):
                statistics = [str(amp_value), str(C), cells, data[amp_value].mean(), data[amp_value].var(),
                              (len(data) + 15) * 105 / (60 * 60), data[amp_value].median()]
                doc = doc.append(pd.DataFrame([statistics], columns=doc_cols), ignore_index=True, sort=False)
                #doc = doc.append(pd.DataFrame([statistics], columns = doc_cols),ignore_index=True)

    if save:
        save_folder = kwargs.get('save_folder', None)
        save_name = kwargs.get('save_name', None)
        book = load_workbook(str(save_folder) + str(save_name)+ '.xlsx')
        writer = pd.ExcelWriter(str(save_folder) + str(save_name)+ '.xlsx', engine='openpyxl')
        writer.book = book
        doc.to_excel(writer, sheet_name='Statistics')
        writer.save()

    return doc

